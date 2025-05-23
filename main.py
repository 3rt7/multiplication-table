import openpyxl, sys, logging
from openpyxl.styles import Font

"""
Create an N multiplication table on an Excel file

USAGE: python3 main.py <N>
<N> - An integer, specifying the dimension of the table

Written by Erfan Zamani <erfanzm99@gmail.com>
"""

# overwrite log-file on every start
logging.basicConfig(level=logging.DEBUG, filename="multiplication-table.log", filemode="w",
                    format="%(levelname)s - %(message)s")

FONT_OBJ = Font(bold=True)

# create nxn table multiplication
table_n = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Multiplication"

for value in range(1, table_n + 1):
    roc = value + 1

    # populate cells
    sheet.cell(row=roc, column=1).value = value
    sheet.cell(row=roc, column=1).font = FONT_OBJ

    sheet.cell(row=1, column=roc).value = value
    sheet.cell(row=1, column=roc).font = FONT_OBJ

    logging.debug(f"current value is {value}:")

    # write multiplication table
    for pos in range(2, roc + 1):
        num_to_multiply = int(sheet.cell(row=pos, column=1).value)
        logging.debug(f"number to multiply {num_to_multiply}")
        result = value*num_to_multiply
        logging.debug(f"result is {result}")

        # keep track of the current row and current column to access
        crow = roc; ccol = pos
        logging.debug(f"cell: ({crow}, {ccol})")
        sheet.cell(row=crow, column=ccol).value = result

        # symmetric value
        crow = pos; ccol = roc
        logging.debug(f"cell: ({crow}, {ccol})")
        sheet.cell(row=crow, column=ccol).value = result


wb.save("multiplication-table.xlsx")
